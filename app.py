# import os

# with open('ip-source.txt') as file:
#    result = file.read().splitlines()

#    for ip in result:
#       os.system('ping -n 2 ' + ip)
import openpyxl as xl
import os

wb = xl.load_workbook('ip.xlsx')
sheet = wb['Sheet1']


def ping(address):
    response = os.system('ping -w 300 ' + address)
    if response == 0:
        pingstatus = "Station Active"
    else:
        pingstatus = "Station Error"

    return pingstatus


for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    if cell.value is not None:
        result = ping(cell.value)
        print(cell.value)
        ping_result = sheet.cell(row, 2)
        ping_result.value = result

wb.save('ip_result.xlsx')









